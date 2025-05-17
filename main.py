from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.spinner import Spinner
from kivy.uix.button import Button
import openpyxl
import os

class KisaanApp(App):
    def build(self):
        layout = BoxLayout(orientation='vertical', padding=20)
        self.month = Spinner(
            text='Baisakh',
            values=('Baisakh','Jestha','Ashadh','Shrawan','Bhadra',
                   'Ashwin','Kartik','Mangsir','Poush','Magh','Falgun','Chaitra')
        )
        btn = Button(text='Generate Reports', size_hint=(1, 0.2))
        btn.bind(on_press=self.generate)
        layout.add_widget(self.month)
        layout.add_widget(btn)
        return layout

    def generate(self, instance):
        base = "/storage/emulated/0/Documents/2080"
        month = self.month.text
        try:
            wb = openpyxl.load_workbook(f"{base}/Management/kisaan.xlsx")
            clients = [wb.active.cell(row=i, column=1).value 
                      for i in range(2, wb.active.max_row + 1)]
            wb.close()
            
            os.makedirs(f"{base}/{month}", exist_ok=True)
            
            for client in clients:
                if client:
                    wb = openpyxl.load_workbook(f"{base}/Management/Template.xlsx")
                    wb.active['E2'] = client
                    wb.active['E3'] = month
                    wb.save(f"{base}/{month}/{client}.xlsx")
                    wb.close()
            
            print(f"✓ Generated {len(clients)} files in 2080/{month}")
        except Exception as e:
            print(f"❌ Error: {str(e)}")

if __name__ == '__main__':
    KisaanApp().run()
